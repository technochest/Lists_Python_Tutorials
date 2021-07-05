import openpyxl

aList = ["John", "Smith", "Adam"]
print(aList)

for name in aList:
    print(name)

len(aList)

aList.reverse()
print(aList)

aList.append("technochest")
print(aList)

if "technochest" in aList:
    print("Found: technochest")

aList.sort()
print(aList)

aList.remove("technochest")
print(aList)

aList.pop()
print(aList)

aList.clear()
print(aList)

aList = [0] * 10
print(aList)

#Combining
bList = ["Adam", "Smith", "John"]
cList = aList + bList
print(aList)
print(bList)
print(cList)

#Slicing
aList.clear()
aList = [10,15,20,63,54,78]
dList = aList[0:3]
print(dList)

#copying list
#newList = aList - we sould not do this because it will create linked duplicate of the list
#linked duplicate means that if you make change in one list then other list will change as well.

copy_aList = aList.copy()

copy_aList = aList[:]
copy_aList.append("Johnny")
print(aList)
print(copy_aList)

#list comprehension
newList = [1, 2, 3, 4, 5 ,6]
listB = [i*5 for i in newList]
print(newList)
print(listB)

wb = openpyxl.load_workbook("C:/Users/JB/PycharmProjects/tutorials/Securities.xlsx", data_only=True)

sht = wb["S&P"]

Sym_List = []
sht_rows = sht.max_row
sht_columns = sht.max_column
for r in range(2,sht_rows+1):
	for c in range(1,2):
		Sym_List.append(sht.cell(r,c).value)

for symbol in Sym_List:
	print(symbol)





